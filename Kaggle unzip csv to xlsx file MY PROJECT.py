import pandas as pd
from openpyxl import workbook, load_workbook
from kaggle.api.kaggle_api_extended import KaggleApi
from openpyxl.utils import get_column_letter

#this program is to unzip a downloaded kaggle .csv file and write it to a new .xlsx file
#must first have a kaggle account generate an API key (MAKE sure it is saved to downloads)


#to generate API key go to settings after being logged in on Kaggle.com and under account scroll down to generate API token

#after generating API key type this into terminal to move it to the right location for program to work(input 1 line at a time)
"""
mkdir -p ~/.kaggle
mv ~/Downloads/kaggle.json ~/.kaggle/
chmod 600 ~/.kaggle/kaggle.json          #ensures proper permissions and security measures for downloaded datasets
"""
#if you are missing libraries type this into the terminal
"""
pip install openpyxl
pip install pandas
pip install kaggle
"""



api = KaggleApi()
api.authenticate()
#can take multiple csv files as a list 
file_path = ['/Applications/Python 3.12/Stock Price Predictor/Project csv files/AKAM_daily_data.csv',
            '/Applications/Python 3.12/Stock Price Predictor/Project csv files/AMZN_daily_data.csv',
            '/Applications/Python 3.12/Stock Price Predictor/Project csv files/CRM_daily_data.csv',
            '/Applications/Python 3.12/Stock Price Predictor/Project csv files/FDX_daily_data.csv',
            '/Applications/Python 3.12/Stock Price Predictor/Project csv files/INTC_daily_data.csv',
            '/Applications/Python 3.12/Stock Price Predictor/Project csv files/LUMN_daily_data.csv',
            '/Applications/Python 3.12/Stock Price Predictor/Project csv files/MA_daily_data.csv',
            '/Applications/Python 3.12/Stock Price Predictor/Project csv files/NFLX_daily_data.csv',
            '/Applications/Python 3.12/Stock Price Predictor/Project csv files/NVDA_daily_data.csv',
            '/Applications/Python 3.12/Stock Price Predictor/Project csv files/PARA_daily_data.csv',
            '/Applications/Python 3.12/Stock Price Predictor/Project csv files/SONY_daily_data.csv',
            '/Applications/Python 3.12/Stock Price Predictor/Project csv files/UPS_daily_data.csv',
            '/Applications/Python 3.12/Stock Price Predictor/Project csv files/V_daily_data.csv',
            '/Applications/Python 3.12/Stock Price Predictor/Project csv files/WBD_daily_data.csv']

#excel file path
new_file_path = '/Applications/Python 3.12/Stock Price Predictor/Project unzipped xlsx files/Stock predictor.xlsx'
#number of sheet names must match number of source files
sheet_datasets = ['Akami Technologies', 'Amazon', 'Salesforce', 'FedEx', 'Intel', 'Lumen Technologies', 'Mastercard', 'Netflix',
           'NVIDIA', 'Paramount Global Class B', 'Sony', 'UPS', 'Visa', 'Warner Bros Discovery']


#uses openpyxl to setup the excel files sheet_datasets and organize the data
excel_file = pd.ExcelFile(new_file_path, engine = 'openpyxl')
sheet_names = excel_file.sheet_names
workbook = load_workbook(new_file_path)
if sheet_names != sheet_datasets:
    for i in range(len(workbook.sheetnames) - 1, -1, -1):
        if workbook.sheetnames[i] not in sheet_datasets:
            workbook.create_sheet(title = sheet_datasets[i], index = i)
    for j in range(len(workbook.sheetnames) - 1, -1, -1):
        if workbook.sheetnames[j] not in sheet_datasets:
            workbook.remove(workbook[workbook.sheetnames[j]])
    workbook.save(new_file_path)


#writes the data to the excel file high runtime complexity but works for a 1 time thing
with pd.ExcelWriter(new_file_path, mode = 'a', engine = "openpyxl", if_sheet_exists = 'replace',) as writer:
    for i in range(0, len(sheet_datasets)):
        df = pd.read_csv(file_path[i])
        df.to_excel(writer, sheet_name = sheet_datasets[i], index = False)
    for sheet_name in sheet_datasets:
        worksheet = writer.sheets[sheet_name] 
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception as e:
                    print(f"Error accessing cell value: {e}")
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
